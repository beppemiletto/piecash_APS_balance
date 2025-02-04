# Definizione della tabella del rendiconto per cassa definito per APS
# con associazione dei conti utilizzati nella contabilità generale (gestionale)
# implementata su GNUCASH per Laboratorio Teatrale di Cambiano

# Struttura della tabella entrate / uscite generale
# Dizionario con campi fissi
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment, Protection
from openpyxl import Workbook
class BalanceTable():
    def __init__(self):
        self.ASSETT = {'COSTO': {'value_n': Decimal(0.00), 'value_n_1': Decimal(0.00)},
                       'DEPREZZAMENTO':{'value_n': Decimal(0.00), 'value_n_1': Decimal(0.00)},
                       'TABLE': {
                           'USCITE': {
                               'title': 'Uscite da investimenti in immobilizzazioni o da deflussi di capitale di terzi',
                               1: {
                                   'DESCRIPTION': '1) Investimenti in immobilizzazioni inerenti alle attività di interesse generale',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': [None]
                               },
                               2: {
                                   'DESCRIPTION': '2) Investimenti in immobilizzazioni inerenti alle attività diverse',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': []
                               },
                               3: {
                                   'DESCRIPTION': '3) Investimenti in attività finanziarie e patrimoniali',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': []
                               },
                               4: {
                                   'DESCRIPTION': '4) Rimborso di finanziamenti per quota capitale e di prestiti',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': []
                               }

                           },

                           'ENTRATE': {
                               'title': 'Entrate da disinvestimenti in immobilizzazioni o da flussi di capitale di terzi',
                               1: {
                                   'DESCRIPTION': '1) Disinvestimenti di immobilizzazioni inerenti alle attività di interesse generale',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': []
                               },
                               2: {
                                   'DESCRIPTION': '2) Disinvestimenti di immobilizzazioni inerenti alle attività diverse',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': []
                               },
                               3: {
                                   'DESCRIPTION': '3) Disinvestimenti di attività finanziarie e patrimoniali',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': []
                               },
                               4: {
                                   'DESCRIPTION': '4) Ricevimento di finanziamenti e di prestiti',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': []
                               },
                               5: {
                                   'DESCRIPTION': '5) Operazioni straordinarie',
                                   'value_n': Decimal(0),
                                   'value_n_1': Decimal(0),
                                   'accounts': ['Entrate:entrataStraordinariaFusione']
                               }

                           },
                           'AV_DIS': {'title': 'Avanzo/disavanzo da entrate e uscite per investimenti e disinvestimenti',
                                      'value_n': Decimal(0),
                                      'value_n_1': Decimal(0)}
                       },
                        'CASSA': {'value_n': Decimal(0.00), 'value_n_1': Decimal(0.00)},
                        'BANCA': {'value_n': Decimal(0.00), 'value_n_1': Decimal(0.00)},
                       'PRESTITI': {'value_n': Decimal(0.00), 'value_n_1': Decimal(0.00)}
                    }
        self.HEADER: str = 'Bilancio esercizio anno {} del Laboratorio Teatrale di Cambiano APS.'
        self.TITLE: str = 'RENDICONTO PER CASSA'
        self.SUBTITLE: str = '''Il  presente prospetto di bilancio è strutturato secondo il MOD. D previsto dal D.L. 5 marzo 2020 del Ministero del Lavoro e delle politiche sociali a titolo “ Adozione della modulistica di bilancio degli enti del Terzo settore.”, secondo quanto previsto dal D.L. 2 agosto 2017 n. 117 e s.m.e.i. ai commi 1 e 2 dell’articolo 13 per ETS con entrate minori di € 22€ 0,000,00.'''
        self.TABLE_HEADER: list = ['USCITE','{n}','{n_1}','ENTRATE','{n}','{n_1}']
        self.TABLE_BODY = {
        # Sezione A - Uscite ed entrate articolo 5 della riforma dello statuto specifica le attività considerate istituzionali
            'A': {
                'USCITE': {'title':'A) Uscite da attività di interesse generale',
                            1: {
                                'DESCRIPTION': '1) Materie prime, sussidiarie, di consumo e di merci',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': ['Uscite:Spese Spettacoli:Scenografie',
                                             'Uscite:Spese Spettacoli:Costumi',
                                             'Uscite:Spese Spettacoli:Trovarobato',
                                             'Uscite:Spese Spettacoli:Consumabili',
                                             'Uscite:Spese Spettacoli:Trucco']
                                },
                            2: {
                                'DESCRIPTION': '2) Servizi ',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': ['Uscite:Pubblicità',
                                             'Uscite:UILT',
                                             'Uscite:Formazione Soci',
                                             'Uscite:DidatticaUscite:PagamentoInsegnanti',
                                             'Uscite:Spese Spettacoli:Rimborsi compagnie ospiti',
                                             'Uscite:EnergiaElettrica',
                                             'Uscite:Gas']
                                },
                            3: {
                                'DESCRIPTION': '3) Godimento beni di terzi',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': ['Uscite:Spese Spettacoli:Locazione Servizi:Affitto sale prova',
                                             'Uscite:Spese Spettacoli:Locazione Servizi:Affitto furgoni',
                                             'Uscite:Spese Spettacoli:Locazione Servizi:Affitto costumi',
                                             'Uscite:LocazioneScuola']
                                },
                            4: {
                                'DESCRIPTION': '4) Personale',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': []
                                },
                            5: {
                                'DESCRIPTION': '5) Uscite diverse di gestione',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': ['Uscite:Benemerenze verso soci',
                                             'Uscite:Rimborso spese soci',
                                             'Uscite:DidatticaUscite:AltriCosti',
                                             'Uscite:Storno per pagamento elettronico errato',
                                             'Uscite:Uscite da pagamenti errati']
                                }

                        },

                'ENTRATE': {
                    'title': 'A) Entrate da attività di interesse generale',
                    1: {
                        'DESCRIPTION': '1) Entrate da quote associative e apporti dei fondatori',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Tesseramento']
                    },
                    2: {
                        'DESCRIPTION': '2) Entrate dagli associati per attività mutuali',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:DidatticaEntrate:CorsoTeatroAdulti',
                                     'Entrate:DidatticaEntrate:CorsoTeatroRagazzi',
                                     'Entrate:DidatticaEntrate:CorsiEstemporaneiStages',
                                     'Entrate:DidatticaEntrate:CorsoDanza',
                                     'Entrate:DidatticaEntrate:corsoGinnastica',
                                     'Entrate:DidatticaEntrate:CorsoMusica',
                                     ]
                    },
                    3: {
                        'DESCRIPTION': '3) Entrate per prestazioni e cessioni ad associati e fondatori',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    4: {
                        'DESCRIPTION': '4) Erogazioni liberali',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Contributi straordinari soci']
                    },
                    5: {
                        'DESCRIPTION': '5) Entrate del 5 per mille',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    6: {
                        'DESCRIPTION': '6) Contributi da soggetti privati',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Contributi da Enti o Privati:Contributi da soggetti privati']
                    },
                    7: {
                        'DESCRIPTION': '7) Entrate per prestazioni e cessioni a terzi',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Fatture Affitto',
                                     'Entrate:Incassi',
                                     'Entrate:Fatture prestazioni']
                    },
                    8: {
                        'DESCRIPTION': '8) Contributi da enti pubblici',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Contributi da Enti o Privati:Contributi da Enti Pubblici']
                    },
                    9: {
                        'DESCRIPTION': '9) Entrate da contratti con enti pubblici',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Contributi da Enti o Privati:Entrate Contratti Enti Pubblici']
                    },
                    10: {
                        'DESCRIPTION': '10) Altre entrate',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Entrata da pagamenti errati']
                    }

                },
                'AV_DIS': {'title': 'Avanzo/disavanzo attività di interesse generale', 'value_n': Decimal(0),'value_n_1': Decimal(0)}
            },
        # Sezione B - Uscite ed entrate fuori da articolo 5 dello statuto che specifica le attività considerate istituzionali
            'B': {
                'USCITE': {
                            'title': 'B) Uscite da attività diverse ',
                            1: {
                                'DESCRIPTION': '1) Materie prime, sussidiarie, di consumo e di merci',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': []
                                },
                            2: {
                                'DESCRIPTION': '2) Servizi ',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': []
                                },
                            3: {
                                'DESCRIPTION': '3) Godimento beni di terzi',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': []
                                },
                            4: {
                                'DESCRIPTION': '4) Personale',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': []
                                },
                            5: {
                                'DESCRIPTION': '5) Uscite diverse di gestione',
                                'value_n': Decimal(0),
                                'value_n_1': Decimal(0),
                                'accounts': []
                                }

                        },

                'ENTRATE': {
                    'title': 'B) Entrate da attività diverse',
                    1: {
                        'DESCRIPTION': '1) Entrate per prestazioni e cessioni ad associati e fondatori',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    2: {
                        'DESCRIPTION': '2) Contributi da soggetti privati',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    3: {
                        'DESCRIPTION': '3) Entrate per prestazioni e cessioni a terzi',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    4: {
                        'DESCRIPTION': '4) Contributi da enti pubblici',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    5: {
                        'DESCRIPTION': '5) Entrate da contratti con enti pubblici',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    6: {
                        'DESCRIPTION': '6) Altre entrate',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    }

                },
                'AV_DIS': {'title': 'Avanzo/disavanzo attività diverse', 'value_n': Decimal(0),'value_n_1': Decimal(0)}

            },
        # Sezione C - Uscite ed entrate articolo 7 raccolta fondi occasionali organizzate occasionali
            'C': {
                'USCITE': {
                    'title': 'C) Uscite da attività di raccolta fondi',
                    1: {
                        'DESCRIPTION': '1) Uscite per raccolte fondi abituali ',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    2: {
                        'DESCRIPTION': '2) Uscite per raccolte fondi occasionali',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    3: {
                        'DESCRIPTION': '3) Altre uscite',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    }
                },

                'ENTRATE': {
                    'title': 'C) Entrate da attività di raccolta fondi',
                    1: {
                        'DESCRIPTION': '1) Entrate per raccolte fondi abituali ',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    2: {
                        'DESCRIPTION': '2) Entrate per raccolte fondi occasionali',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    3: {
                        'DESCRIPTION': '3) Altre entrate',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    }

                },
                'AV_DIS': {'title': 'Avanzo/disavanzo attività di raccolta fondi', 'value_n': Decimal(0), 'value_n_1': Decimal(0)}

            },
            'D': {
                'USCITE': {
                    'title': 'D) Uscite da attività finanziarie e patrimoniali',
                    1: {
                        'DESCRIPTION': '1) Su rapporti bancari ',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Uscite:Spese banca']
                    },
                    2: {
                        'DESCRIPTION': '2) Su investimenti finanziari ',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    3: {
                        'DESCRIPTION': '3) Su patrimonio edilizio ',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    4: {
                        'DESCRIPTION': '4) Su altri beni patrimoniali ',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    5: {
                        'DESCRIPTION': '5) Altre uscite',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    }

                },

                'ENTRATE': {
                    'title': 'D) Entrate da attività finanziarie e patrimoniali',
                    1: {
                        'DESCRIPTION': '1) Da rapporti bancari',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Interessi bancari a credito']
                    },
                    2: {
                        'DESCRIPTION': '2) Da altri investimenti finanziari',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    3: {
                        'DESCRIPTION': '3) Da patrimonio edilizio',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    4: {
                        'DESCRIPTION': '4) Da altri beni patrimoniali',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    5: {
                        'DESCRIPTION': '5) Altre entrate',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    }

                },
                'AV_DIS': {'title': 'Avanzo/disavanzo attività finanziarie e patrimoniali', 'value_n': Decimal(0), 'value_n_1': Decimal(0)}

            },
            'E': {
                'USCITE': {
                    'title': 'E) Uscite di supporto generale',
                    1: {
                        'DESCRIPTION': '1) Materie prime, sussidiarie, di consumo e di merci',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Uscite:Apparecchiature:Luci',
                                     'Uscite:Apparecchiature:Audio',
                                     'Uscite:Apparecchiature:Apparecchiature generiche',
                                     'Uscite:Pulizie:Materiale',
                                     'Uscite:Pulizie:Attrezzature',
                                     'Uscite:Cancelleria',
                                     'Uscite:Spese Varie'
                                     # 'Uscite:Spese Varie:BevandeCibo'
                                     ]
                    },
                    2: {
                        'DESCRIPTION': '2) Servizi ',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Uscite:Telefonia',
                                     'Uscite:Assicurazioni',
                                     'Uscite:SIAE-Diritti',
                                     'Uscite:Spese Spettacoli:Riprese Video',
                                     'Uscite:Pulizie:Personale',
                                     'Uscite:Segreteria Gestione',
                                     'Uscite:CONTABILITA  MANODOPERA',
                                     'Uscite:DidatticaUscite:segreteria']
                    },
                    3: {
                        'DESCRIPTION': '3) Godimento beni di terzi',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    4: {
                        'DESCRIPTION': '4) Personale',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    5: {
                        'DESCRIPTION': '5) Altre uscite',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': [
                                     'Uscite:manutenzioneLocali:ManutenzioneScuola',
                                     'Uscite:manutenzioneLocali:Manutenzione Teatro',
                                     'Uscite:Tasse:Imposta di bollo su rendiconto banca',
                                     'Uscite:Tasse:IVA non deducibile',
                                     'Uscite:Tasse:TARI',
                                     'Uscite:Tasse:Sanzioni amministrative',
                                     'Uscite:Tasse:AgenziaEntrate',
                                     'Uscite:Tasse:IVA',
                                     'Uscite:Tasse:Spese registrazione Atti Marche da Bollo',
                                     'Uscite:Tasse:IRES e IRAP'],

                    }

                },

                'ENTRATE': {
                    'title': 'E) Entrate di supporto generale',
                    1: {
                        'DESCRIPTION': '1) Entrate da distacco del personale',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    2: {
                        'DESCRIPTION': '2) Altre entrate di supporto generale',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:Introiti distributori']
                    }
                }

            },
            'F': {
                'USCITE': {
                    'title': 'Uscite da investimenti in immobilizzazioni o da deflussi di capitale di terzi',
                    1: {
                        'DESCRIPTION': '1) Investimenti in immobilizzazioni inerenti alle attività di interesse generale',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    2: {
                        'DESCRIPTION': '2) Investimenti in immobilizzazioni inerenti alle attività diverse',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    3: {
                        'DESCRIPTION': '3) Investimenti in attività finanziarie e patrimoniali',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    4: {
                        'DESCRIPTION': '4) Rimborso di finanziamenti per quota capitale e di prestiti',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    }

                },

                'ENTRATE': {
                    'title': 'Entrate da disinvestimenti in immobilizzazioni o da flussi di capitale di terzi',
                    1: {
                        'DESCRIPTION': '1) Disinvestimenti di immobilizzazioni inerenti alle attività di interesse generale',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    2: {
                        'DESCRIPTION': '2) Disinvestimenti di immobilizzazioni inerenti alle attività diverse',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },
                    3: {
                        'DESCRIPTION': '3) Disinvestimenti di attività finanziarie e patrimoniali',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': []
                    },

                    4: {
                        'DESCRIPTION': '4) Ricevimento di finanziamenti e di prestiti',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Passività:Anticipi spese da soci']
                    },
                    5: {
                        'DESCRIPTION': '5) Operazioni straordinarie',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Entrate:entrataStraordinariaFusione']
                    },
                }

            },
            'Patrimonio': {

                'DARE': {
                    'title': 'Cassa e banca',
                    1: {
                        'DESCRIPTION': 'Cassa',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Attività:Cassa']
                    },
                    2: {
                        'DESCRIPTION': 'Depositi bancari e postali',
                        'value_n': Decimal(0),
                        'value_n_1': Decimal(0),
                        'accounts': ['Attività:BANCA:BancaASTI',
                                     'Attività:BANCA:UNICREDIT']
                    }
                }
            }

        }
        self.BALANCE = {'GTU': {'value_n': Decimal(0.00), 'value_n_1': Decimal(0.00)},
                        'GTE': {'value_n': Decimal(0.00), 'value_n_1': Decimal(0.00)},
                        'surplus_Balance': {'value_n': Decimal(0.00), 'value_n_1': Decimal(0.00)}
                        }
        self.PERIODS = {'period_n': {'begin': None, 'end': None}, 'period_n_1': {'begin': None, 'end': None}}
        self.GNUCASH_FILE = None

    def save_balance(self, filename=None):
        pass

class ExcelBalanceTable():

    def __init__(self, fname=None):
        self.ExcelFileName = fname
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "Bilancio"
        self.ws1.column_dimensions['A'].width = 40
        self.ws1.column_dimensions['B'].width = 12
        self.ws1.column_dimensions['C'].width = 12
        self.ws1.column_dimensions['D'].width = 40
        self.ws1.column_dimensions['E'].width = 12
        self.ws1.column_dimensions['F'].width = 12
        self.cp: int = 1
        self.rp: int = 1
        self.currency_format = '€ ###.##0,00'

    def writeline(self, dataline=None, row_height=None, color=None, fontsize=None, bold=None, italic=None, halign=None,
                  wrap=None, border=None, euro=None):
        curr_cp = self.cp
        curr_rp = self.rp
        if dataline[0] is not None:
            curr_cp = dataline[0]
        if dataline[1] is not None:
            curr_rp = dataline[1]
        for val in dataline[2]:
            self.cell = self.ws1.cell(column=curr_cp, row=curr_rp, value=val)
            # if euro is not None:
            #     self.cell.number_format = self.currency_format
            if dataline[3] is not None:
                self.ws1.merge_cells(start_row=curr_rp, start_column=curr_cp, end_row=curr_rp, end_column=dataline[3])
            if border is not None:
                self.cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                right=Side(border_style='thin', color='FF000000'),
                                top=Side(border_style='thin', color='FF000000'),
                                bottom=Side(border_style='thin', color='FF000000'))

            self.cell.font = Font(size=fontsize, bold=bold, italic=italic)
            self.cell.alignment = Alignment(horizontal=halign, wrap_text=wrap)
            if row_height is not None:
                self.ws1.row_dimensions[curr_rp].height = row_height

            curr_cp += 1

        return self.cell




    def save(self):
        self.wb.save(filename=self.ExcelFileName)



from decimal import Decimal